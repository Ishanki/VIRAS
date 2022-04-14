# -*- coding: utf-8 -*-
"""
Created on Tue May 18 10:21:59 2021
@author: Ishanki
Python version 3.7.6

N.B. Close contact/transfers with friend's surfaces currently commented out

"""

import openpyxl as wb
import random as random

# =============================================================================
# Inputs
# =============================================================================
# PATH ='/Users/Ishanki/OneDrive/Documents/Python Scripts/Data_nan_new/'
NAME = 'Inputs.xlsx'

STUDENTS = 39
INFECTED = [1]  #Enter student ID of infected individual(s)

# Inputs for Private surfaces
PRIVATE_OBJ = ['Desk', 'Chair']
PRIV_AREAS = [6000, 4260]  # Areas of private objs
PRIV_CA = [73.5, 73.5]  # Contact Areas of private objs
PRIV_MATERIAL = ['Non-porous', 'Porous']  # Material types of private 
PRIV_CF = [20.2, 8.9]  # Contact freq
## Reduce chair closetime
PRIV_CLOSETIME = [0.8, 0.4]  # % time spent with private surfaces
PRIV_TRANSFER = [0.7, 0.05]  # % Large droplets transferred
HANDS_TRANSFER = 0.1 # Percentage of large droplets transferred to own hand

# Inputs for public surfaces
PUBLIC_OBJ = ['Cabinethandle', 'Printerscreen', 'WDbutton', 'Doorhandle']  # Leave empty list if no public objects
PUBLIC_NOS = [3, 1, 1, 1]
PUBL_AREAS = [10, 35, 12, 100]
PUBL_CA = [7, 17.5, 6, 70]
PUBL_MATERIAL = ['Non-porous', 'Non-porous', 'Non-porous', 'Steel']
PUBL_CF = [0.14, 0.28, 0.31, 0.05]  # Contact freqs
# If students are assigned cabinets, state which students use which cabinets
PROXIMITY = {'Cabinethandle_1': list(range(1,14)), 
             'Cabinethandle_2': list(range(14,27)),
             'Cabinethandle_3': list(range(27,40)),
             # 'Printer_1': list(range(1,20)),
             # 'Printer_2': list(range(20,40)),
             }
PUBL_CLOSETIME = [0.03, 0.01, 0.005, 0.005]
PUBL_TRANSFER = [0.01, 0.05, 0.05, 0.01]

STUDENT_GROUPS = 3 # Enter number of friends in groups, 0 if no groups
GROUPS_CLOSETIME = 0.125  # Fraction of time spent with members of group
GROUPS_TRANSFER = [0.01, 0.005]  # [hands, mucous]

# Change these if you want to
TIME_IN = 0 
DURATION = 8  # TODO: should these be different for each individual?
LRV_HANDS = 1.10
LRV_MUCOSA = 0
HMRATE = 0.35
MHRATE = 0.5
HL_HANDS = 3.5
HL_MUCOSA = 10000000000
M_CA = 7.67
M_CF = 16
H_AREA = 147
M_AREA = 391.65
DOSE_R = 395000
INF_SHEDDING_R = 11300000
INH_R = 0.39
LDR = 0.5
DEP_H = 0
DEP_M = 0
INF_MUC_CONT = 4000000
CLEAN_EFF = 1

# =============================================================================
# Grouping individuals
# =============================================================================
# Just making sure I get the same 'random' result everytime
random.seed(2) 

student_list=[]
for i in range(1,STUDENTS+1):
    student_list.append(i)

# random.shuffle(student_list)
# print(student_list)
def chunks(student_list, n):
    """Yield successive n-sized chunks from list."""
    for i in range(0, len(student_list), n):
        yield student_list[i:i + n]

groups_dict = {}
if STUDENT_GROUPS !=0:
    groups = list(chunks(student_list, STUDENT_GROUPS))
    for i in range(1,len(groups)+1):
        groups_dict[i] = groups[i-1]
# print(groups_dict)

# =============================================================================
# Changing spreadsheet
# =============================================================================
file = wb.load_workbook(filename=NAME, data_only=True)
Objects = file['Objects']
People = file['People']
Contacts = file['Contacts']
CloseTime = file['CloseTime']
CloseTransfer = file['CloseTransfer']


# Adding objects to the 'Objects' sheet
count=0
obj_list = []
person_obj = {}
for i in range(1,STUDENTS+1):
    personal_obj = []
    for j in range(1,(len(PRIVATE_OBJ)+1)):
        count +=1
        # print(count)
        # Changing ID
        Objects.cell(row=count+1,column=1).value = count
        # Changing Object params
        Objects.cell(row=count+1,column=2).value = PRIVATE_OBJ[j-1]+'_'+str(i)
        Objects.cell(row=count+1,column=3).value = PRIV_AREAS[j-1]
        Objects.cell(row=count+1,column=4).value = PRIV_CA[j-1]
        Objects.cell(row=count+1,column=5).value = PRIV_MATERIAL[j-1]
        Objects.cell(row=count+1,column=6).value = 0  # Contam
        Objects.cell(row=count+1,column=7).value = 0  # Dep rate
        Objects.cell(row=count+1,column=8).value = 1  # Cleaning eff
        obj_list.append(PRIVATE_OBJ[j-1]+'_'+str(i))
        personal_obj.append(PRIVATE_OBJ[j-1]+'_'+str(i))
    person_obj[i] = personal_obj
 
if PUBLIC_OBJ:
    for k in range(1,len(PUBLIC_OBJ)+1):
        for n in range(1,PUBLIC_NOS[k-1]+1):
            count += 1
            # print(count)
            Objects.cell(row=count+1,column=1).value = count
            if PUBLIC_NOS[k-1] > 1:
                Objects.cell(row=count+1,column=2).value = PUBLIC_OBJ[k-1]+'_'+str(n)
            else:
                Objects.cell(row=count+1,column=2).value = PUBLIC_OBJ[k-1]
            Objects.cell(row=count+1,column=3).value = PUBL_AREAS[k-1]
            Objects.cell(row=count+1,column=4).value = PUBL_CA[k-1]
            Objects.cell(row=count+1,column=5).value = PUBL_MATERIAL[k-1]
            Objects.cell(row=count+1,column=6).value = 0  # Contam
            Objects.cell(row=count+1,column=7).value = 0  # Dep rate
            Objects.cell(row=count+1,column=8).value = 1  # Cleaning eff
            obj_list.append(PUBLIC_OBJ[k-1]+'_'+str(n))
else:
    print('No public objects')
      
# Adding students to 'People' sheet 
people_list = []
for i in range(1,STUDENTS+1):
    people_list.append(i)
    People.cell(row=i+1,column=1).value = i
    People.cell(row=i+1,column=2).value = TIME_IN
    People.cell(row=i+1,column=3).value = DURATION
    People.cell(row=i+1,column=4).value = LRV_HANDS
    People.cell(row=i+1,column=5).value = LRV_MUCOSA
    People.cell(row=i+1,column=6).value = HMRATE
    People.cell(row=i+1,column=7).value = MHRATE
    People.cell(row=i+1,column=8).value = HL_HANDS
    People.cell(row=i+1,column=9).value = HL_MUCOSA
    People.cell(row=i+1,column=10).value = M_CA
    People.cell(row=i+1,column=11).value = M_CF
    People.cell(row=i+1,column=12).value = H_AREA
    People.cell(row=i+1,column=13).value = M_AREA
    People.cell(row=i+1,column=14).value = DOSE_R
    People.cell(row=i+1,column=15).value = 0
    People.cell(row=i+1,column=16).value = INH_R
    People.cell(row=i+1,column=17).value = LDR
    People.cell(row=i+1,column=18).value = DEP_H
    People.cell(row=i+1,column=19).value = DEP_M
    People.cell(row=i+1,column=20).value = 0
    People.cell(row=i+1,column=21).value = 0
    People.cell(row=i+1,column=22).value = CLEAN_EFF
    
    # If individual is infected, then some parameters are updated
    for n in range(len(INFECTED)):
        if i == INFECTED[n]:
            # print(i)
            People.cell(row=i+1,column=15).value = INF_SHEDDING_R
            People.cell(row=i+1,column=20).value = INF_MUC_CONT
            People.cell(row=i+1,column=21).value = 1
            
    
# print(obj_list)
# print(people_list)
# print(person_obj)
contact_table = {}
for i in people_list:
    for j in range(1,len(obj_list)+1):
        # creating dict of zeros
        contact_table[i,j] = 0
        # checking personal objects and changing value accordingly
        for k, v in person_obj.items():
            # print(v)
            if i == k:
                if obj_list[j-1] in v:
                    for obj in range(len(PRIVATE_OBJ)):
                        if str(PRIVATE_OBJ[obj]) in obj_list[j-1]:
                            contact_table[i,j] = PRIV_CF[obj]
        
        # checking public objects and changing values accordingly
        if PUBLIC_OBJ:
            for item in range(len(PUBLIC_OBJ)):
                if str(PUBLIC_OBJ[item]) in obj_list[j-1]:
                    if PUBLIC_NOS[item] > 1:
                        if PROXIMITY:
                            for k, v in PROXIMITY.items():
                                if k == obj_list[j-1]:
                                    # print(k)
                                    if i in v:
                                        contact_table[i,j] = PUBL_CF[item]
                                        # print(PUBL_CF[item])
                                    else:
                                        contact_table[i,j] = 0
                    else:
                        contact_table[i,j] = PUBL_CF[item]
                    # print(i,j)

# print(contact_table)    
# Contact frequencies to Excel
# N.B. rows are objects, columns are people, so I've done the swap below
for (i,j), v in contact_table.items():      
    Contacts.cell(row=j+1, column=i).value = v

own_body = {}
for i in people_list:
    hands = 'Hands_'+str(i)
    obj_list.append(hands)
    own_body[i] = []
    own_body[i].append(hands)
for i in people_list:
    muc = 'Mucosa_'+str(i)
    obj_list.append(muc)  
    own_body[i].append(muc)
# print(obj_list)
# print(own_body)

closetime_table = {}
closetransfer_table = {}
for j in range(1,len(obj_list)+1):
    for i in people_list:
        closetime_table[j,i] = 0
        closetransfer_table[j,i] = 0
        # Checking for private objects and assigning value
        for k, v in person_obj.items():
            if i == k:
                if obj_list[j-1] in v:
                    for obj in range(len(PRIVATE_OBJ)):
                        if str(PRIVATE_OBJ[obj]) in obj_list[j-1]:
                            closetime_table[j,i] = PRIV_CLOSETIME[obj]
                            closetransfer_table[j,i] = PRIV_TRANSFER[obj]
        
        # checking for own hand and mucous membrane                   
        for k, v in own_body.items():
            if i == k:
                if obj_list[j-1] in v:
                    closetime_table[j,i] = 1
                    if 'Hand' in obj_list[j-1]:
                        closetransfer_table[j,i] = HANDS_TRANSFER
                        
                
        # Checking for groups of friends
        for key, val in groups_dict.items():
            if i in val:
                friends = [v for v in val if v!=i]
                for f in friends:
                    ## If close contact with friend's surfaces
                    # for k, v in person_obj.items():
                    #     if f == k:
                    #         if obj_list[j-1] in v:
                    #             closetime_table[j,i] = GROUPS_CLOSETIME
                    ## If close transfer to friend's surfaces
                    # for k, v in person_obj.items():
                    #     if f == k:
                    #         if obj_list[j-1] in v:
                    #             closetransfer_table[j,i] = GROUPS_TRANSFER[0]
                    # Close contact with friend's hands and mucosa
                    for k, v in own_body.items():
                        if f == k:
                            if obj_list[j-1] in v:
                                closetime_table[j,i] = GROUPS_CLOSETIME
                                if 'Hand' in obj_list[j-1]:
                                    closetransfer_table[j,i] = GROUPS_TRANSFER[0]
                                elif 'Mucosa' in obj_list[j-1]:
                                    closetransfer_table[j,i] = GROUPS_TRANSFER[1]
                    
        
        # checking for public surfaces and assigning value
        if PUBLIC_OBJ:
            for item in range(len(PUBLIC_OBJ)):
                if str(PUBLIC_OBJ[item]) in obj_list[j-1]:
                    if PUBLIC_NOS[item] > 1:
                        if PROXIMITY:
                            for k, v in PROXIMITY.items():
                                if k == obj_list[j-1]:
                                    # print(k)
                                    if i in v:
                                        closetime_table[j,i] = PUBL_CLOSETIME[item]
                                        closetransfer_table[j,i] = PUBL_TRANSFER[item]
                                        # print(PUBL_CF[item])
                                    else:
                                        closetime_table[j,i] = 0
                                        closetransfer_table[j,i] = 0
                    else:
                        closetime_table[j,i] = PUBL_CLOSETIME[item]
                        closetransfer_table[j,i] = PUBL_TRANSFER[item]

# print(closetime_table)
# Close contact time dictionary to Excel
for (j,i), v in closetime_table.items():      
    CloseTime.cell(row=j+1, column=i).value = v

# Close contact transfer dictionary to Excel        
for (j,i), v in closetransfer_table.items():      
    CloseTransfer.cell(row=j+1, column=i).value = v   

# Saving and closing excel file
file.save(filename=NAME)

print('All done.')
    

                
                



